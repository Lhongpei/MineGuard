"""Safe file parsing and deterministic five-quantity normalisation suggestions."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import re
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import PurePosixPath
from typing import Any
from zoneinfo import ZoneInfo

from .errors import ImportContentError
from .five_quantity_exchange import MineIdentity
from .util import jcs_json, utc_text

MAX_IMPORT_BYTES = 20 * 1024 * 1024
MAX_SHEETS = 32
MAX_ROWS = 10000
MAX_COLUMNS = 256
MAX_CELLS = 500000
MAX_ZIP_ENTRIES = 5000
MAX_ZIP_UNCOMPRESSED = 80 * 1024 * 1024
ALLOWED_SUFFIXES = {".et", ".xls", ".xlsx", ".csv", ".json", ".jsonl"}
METRICS = (
    "ventilation_m3_min",
    "electricity_kwh",
    "detonators_count",
    "explosives_kg",
    "mine_entry_persons",
    "production_t",
)
SHIFT_KEYS = ("zero_shift", "eight_shift", "four_shift")
PERIOD_KEYS = ("daily_total", *SHIFT_KEYS)
UNITS = {
    "ventilation_m3_min": "m3/min",
    "electricity_kwh": "kWh",
    "detonators_count": "count",
    "explosives_kg": "kg",
    "mine_entry_persons": "person",
    "production_t": "t",
}
AGGREGATIONS = {
    "ventilation_m3_min": "time_weighted_average",
    "electricity_kwh": "sum",
    "detonators_count": "sum",
    "explosives_kg": "sum",
    "mine_entry_persons": "sum",
    "production_t": "sum",
}
_VALUE_UNIT_SUFFIXES = {
    "ventilation_m3_min": ("m3/min", "m³/min", "立方米/分钟", "立方米每分钟"),
    "electricity_kwh": ("kwh", "千瓦时", "度"),
    "detonators_count": ("发", "枚", "个"),
    "explosives_kg": ("kg", "千克", "公斤"),
    "mine_entry_persons": ("人次", "人"),
    "production_t": ("t", "吨"),
}
_DATE_ALIASES = {"日期", "date", "统计日期"}
_INFERRED_DATE_HEADER_HINTS = (
    "businessdate",
    "reportdate",
    "bizdate",
    "date",
    "day",
    "业务日",
    "报表日",
    "数据日",
    "核算日",
    "统计日",
)
# Keep the legacy aliases readable at the import boundary.  All generated
# payloads use mine_entry_persons, so the old wording can never leak back onto
# the V2 wire.
_METRIC_ALIASES = (
    ("detonators_count", "detonators_count", False),
    ("数码电子雷管", "detonators_count", False),
    ("电子雷管", "detonators_count", False),
    ("工业雷管", "detonators_count", False),
    ("雷管", "detonators_count", False),
    ("explosives_kg", "explosives_kg", False),
    ("工业炸药", "explosives_kg", False),
    ("乳化炸药", "explosives_kg", False),
    ("炸药", "explosives_kg", False),
    ("mine_entry_persons", "mine_entry_persons", False),
    ("underground_person_entries", "mine_entry_persons", False),
    ("入井人员量", "mine_entry_persons", False),
    ("下井人员量", "mine_entry_persons", False),
    ("入井人员", "mine_entry_persons", False),
    ("下井人员", "mine_entry_persons", False),
    ("入井人数", "mine_entry_persons", False),
    ("下井人数", "mine_entry_persons", False),
    ("入井人次", "mine_entry_persons", False),
    ("下井人次", "mine_entry_persons", False),
    ("labor_persons", "mine_entry_persons", True),
    ("用工量", "mine_entry_persons", True),
    ("用工", "mine_entry_persons", True),
    ("人数", "mine_entry_persons", True),
    ("ventilation_m3_min", "ventilation_m3_min", False),
    ("wind_m3_min", "ventilation_m3_min", False),
    ("通风量", "ventilation_m3_min", False),
    ("风量", "ventilation_m3_min", False),
    ("ventilation", "ventilation_m3_min", False),
    ("wind", "ventilation_m3_min", False),
    ("electricity_kwh", "electricity_kwh", False),
    ("耗电量", "electricity_kwh", False),
    ("用电量", "electricity_kwh", False),
    ("电量", "electricity_kwh", False),
    ("electricity", "electricity_kwh", False),
    ("production_t", "production_t", False),
    ("原煤产量", "production_t", False),
    ("产量", "production_t", False),
    ("production", "production_t", False),
    ("爆破器材量", "fire_material", False),
    ("民爆物品量", "fire_material", False),
    ("火工品量", "fire_material", False),
    ("火工品", "fire_material", False),
)
_SHIFT_ALIASES = {
    "zero_shift": "zero_shift",
    "零点班": "zero_shift",
    "00点班": "zero_shift",
    "0点班": "zero_shift",
    "zero": "zero_shift",
    "eight_shift": "eight_shift",
    "八点班": "eight_shift",
    "08点班": "eight_shift",
    "8点班": "eight_shift",
    "eight": "eight_shift",
    "four_shift": "four_shift",
    "四点班": "four_shift",
    "16点班": "four_shift",
    "4点班": "four_shift",
    "four": "four_shift",
    "daily_total": "daily_total",
    "日统计": "daily_total",
    "日合计": "daily_total",
    "合计": "daily_total",
    "daily": "daily_total",
}
_FIRE_NUMBER = re.compile(
    r"(?:(?:数码电子|电子|工业|煤矿许用|乳化|水胶|铵油|粉状)\s*)?"
    r"(雷管|炸药)\s*[:：=]?\s*(-?(?:[0-9]+(?:\.[0-9]+)?|\.[0-9]+))"
)
_UNSUPPORTED_FIRE_COMPONENT_ALIASES = ("导爆索", "导爆管", "起爆具")
_EXPLICIT_MISSING_VALUES = {"", "-", "—", "/", "无", "缺失"}
_FORMULA_PREFIXES = ("=", "+", "@")
_GROUPED_NUMBER = re.compile(r"^[0-9]{1,3}(?:,[0-9]{3})+(?:\.[0-9]+)?$")
_HEADER_NUMBER_WITH_UNIT = re.compile(
    r"^\s*[+-]?(?:[0-9]+(?:[,.][0-9]+)*|\.[0-9]+)\s*"
    r"(?:m3/min|m³/min|kwh|kg|t|立方米/分钟|千瓦时|千克|公斤|"
    r"人次|吨|度|人|发|枚|个)\s*$",
    flags=re.IGNORECASE,
)


@dataclass(frozen=True)
class ParsedSheet:
    name: str
    rows: list[list[Any]]


@dataclass(frozen=True)
class TableLayout:
    header_start: int
    data_start: int
    date_column: int
    width: int
    normalized_headers: tuple[str, ...]
    display_headers: tuple[str, ...]
    date_inferred: bool


def _strict_json_loads(value: str) -> Any:
    def reject_duplicates(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
        result: dict[str, Any] = {}
        for key, item in pairs:
            if key in result:
                raise ImportContentError(f"JSON 包含重复字段：{key}")
            result[key] = item
        return result

    def reject_constant(value: str) -> None:
        raise ImportContentError(f"JSON 包含非标准数值：{value}")

    try:
        parsed = json.loads(
            value,
            object_pairs_hook=reject_duplicates,
            parse_constant=reject_constant,
        )
    except ImportContentError:
        raise
    except (RecursionError, ValueError) as error:
        raise ImportContentError("JSON 结构非法或嵌套层级过深") from error
    stack: list[tuple[Any, int]] = [(parsed, 1)]
    node_count = 0
    while stack:
        item, depth = stack.pop()
        node_count += 1
        if node_count > 500_000:
            raise ImportContentError("JSON 结构节点过多")
        if depth > 64:
            raise ImportContentError("JSON 嵌套层级不能超过 64")
        if isinstance(item, dict):
            stack.extend((child, depth + 1) for child in item.values())
        elif isinstance(item, list):
            stack.extend((child, depth + 1) for child in item)
    return parsed


def _safe_filename(value: str) -> str:
    if not isinstance(value, str):
        raise ImportContentError("文件名必须是字符串")
    name = value.strip()
    if (
        not name
        or len(name) > 255
        or name in {".", ".."}
        or any(character in name for character in "/\\")
        or any(ord(character) < 32 or ord(character) == 127 for character in name)
    ):
        raise ImportContentError("文件名非法")
    suffix = "." + name.rsplit(".", 1)[-1].lower() if "." in name else ""
    if suffix not in ALLOWED_SUFFIXES:
        raise ImportContentError("仅支持 ET、XLS、XLSX、CSV 和 JSON")
    return name


def _check_bytes(filename: str, content: bytes) -> str:
    name = _safe_filename(filename)
    if not isinstance(content, bytes) or not content:
        raise ImportContentError("导入文件不能为空")
    if len(content) > MAX_IMPORT_BYTES:
        raise ImportContentError("单个导入文件不能超过 20 MiB")
    return name


def _check_xlsx_archive(content: bytes) -> None:
    try:
        archive = zipfile.ZipFile(io.BytesIO(content))
    except (OSError, zipfile.BadZipFile) as error:
        raise ImportContentError("XLSX/ET 压缩容器损坏") from error
    with archive:
        entries = archive.infolist()
        if len(entries) > MAX_ZIP_ENTRIES:
            raise ImportContentError("工作簿压缩条目过多")
        total = 0
        for entry in entries:
            path = PurePosixPath(entry.filename)
            if path.is_absolute() or ".." in path.parts:
                raise ImportContentError("工作簿包含不安全路径")
            total += entry.file_size
            if total > MAX_ZIP_UNCOMPRESSED:
                raise ImportContentError("工作簿解压后超过 80 MiB")
            if entry.compress_size == 0 and entry.file_size > 0:
                raise ImportContentError("工作簿包含异常压缩条目")
            if entry.compress_size and entry.file_size / entry.compress_size > 200:
                raise ImportContentError("工作簿疑似压缩炸弹")


def _xlsx_sheets(content: bytes) -> list[ParsedSheet]:
    _check_xlsx_archive(content)
    try:
        import openpyxl
    except ImportError as error:  # pragma: no cover - dependency guard
        raise ImportContentError("缺少 XLSX 解析组件 openpyxl") from error
    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as error:
        raise ImportContentError("无法读取 XLSX/ET 工作簿") from error
    try:
        if len(workbook.sheetnames) > MAX_SHEETS:
            raise ImportContentError("工作簿 sheet 数量过多")
        result: list[ParsedSheet] = []
        cell_count = 0
        for worksheet in workbook.worksheets:
            rows: list[list[Any]] = []
            for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
                if row_index >= MAX_ROWS:
                    raise ImportContentError("工作簿行数过多")
                values = list(row[:MAX_COLUMNS])
                if len(row) > MAX_COLUMNS:
                    raise ImportContentError("工作簿列数过多")
                cell_count += len(values)
                if cell_count > MAX_CELLS:
                    raise ImportContentError("工作簿单元格数量过多")
                rows.append(values)
            if rows:
                result.append(ParsedSheet(str(worksheet.title)[:128], rows))
        return result
    finally:
        workbook.close()


def _xls_sheets(content: bytes) -> list[ParsedSheet]:
    try:
        import xlrd
    except ImportError as error:  # pragma: no cover - dependency guard
        raise ImportContentError("缺少 ET/XLS 解析组件 xlrd") from error
    try:
        workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
    except Exception as error:
        raise ImportContentError("无法读取 ET/XLS 工作簿") from error
    try:
        if workbook.nsheets > MAX_SHEETS:
            raise ImportContentError("工作簿 sheet 数量过多")
        result: list[ParsedSheet] = []
        cell_count = 0
        for worksheet in workbook.sheets():
            if worksheet.nrows > MAX_ROWS or worksheet.ncols > MAX_COLUMNS:
                raise ImportContentError("工作簿行列数超过安全限制")
            cell_count += worksheet.nrows * worksheet.ncols
            if cell_count > MAX_CELLS:
                raise ImportContentError("工作簿单元格数量过多")
            if worksheet.nrows:
                result.append(
                    ParsedSheet(
                        str(worksheet.name)[:128],
                        [
                            worksheet.row_values(index)
                            for index in range(worksheet.nrows)
                        ],
                    )
                )
        return result
    finally:
        workbook.release_resources()


def _decode_text_with_encoding(content: bytes) -> tuple[str, str]:
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            label = (
                "utf-8-sig"
                if encoding == "utf-8-sig" and content.startswith(b"\xef\xbb\xbf")
                else "utf-8"
                if encoding == "utf-8-sig"
                else encoding
            )
            return content.decode(encoding), label
        except UnicodeDecodeError:
            continue
    raise ImportContentError("文本文件必须是 UTF-8 或 GB18030")


def _decode_text(content: bytes) -> str:
    return _decode_text_with_encoding(content)[0]


def _csv_rows(content: bytes) -> tuple[list[list[str]], str, str]:
    text, encoding = _decode_text_with_encoding(content)
    if "\x00" in text:
        raise ImportContentError("CSV 包含 NUL 控制字符")
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel
    try:
        rows = [
            list(row)
            for row in csv.reader(io.StringIO(text), dialect, strict=True)
        ]
    except csv.Error as error:
        raise ImportContentError("CSV 引号、分隔符或字段长度格式非法") from error
    if len(rows) > MAX_ROWS or any(len(row) > MAX_COLUMNS for row in rows):
        raise ImportContentError("CSV 行列数超过安全限制")
    if sum(len(row) for row in rows) > MAX_CELLS:
        raise ImportContentError("CSV 单元格数量超过安全限制")
    return rows, encoding, str(dialect.delimiter)


def _csv_sheets(content: bytes) -> list[ParsedSheet]:
    rows, _encoding, _delimiter = _csv_rows(content)
    return [ParsedSheet("CSV", rows)]


def _normal_text(value: Any) -> str:
    return str(value).strip().replace(" ", "").replace("\u3000", "").casefold()


def _number(value: Any, *, integer: bool = False) -> int | float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        parsed = float(value)
    elif isinstance(value, str):
        clean = value.strip()
        if clean in _EXPLICIT_MISSING_VALUES:
            return None
        if clean.startswith(_FORMULA_PREFIXES):
            return None
        if "," in clean and _GROUPED_NUMBER.fullmatch(clean) is None:
            return None
        clean = clean.replace(",", "")
        try:
            parsed = float(clean)
        except ValueError:
            return None
    else:
        return None
    if not math.isfinite(parsed) or parsed < 0 or parsed > 1_000_000_000_000_000:
        return None
    if integer:
        return int(parsed) if parsed.is_integer() else None
    return int(parsed) if parsed.is_integer() else parsed


def _metric_number(metric: str, value: Any) -> int | float | None:
    candidate = value
    if isinstance(value, str):
        clean = value.strip()
        for suffix in _VALUE_UNIT_SUFFIXES[metric]:
            if clean.casefold().endswith(suffix.casefold()):
                clean = clean[: -len(suffix)].strip()
                break
        candidate = clean
    return _number(
        candidate,
        integer=metric in {"detonators_count", "mine_entry_persons"},
    )


def _explicit_missing(value: Any) -> bool:
    return value is None or (
        isinstance(value, str) and value.strip() in _EXPLICIT_MISSING_VALUES
    )


def _formula_like(value: Any) -> bool:
    return isinstance(value, str) and value.lstrip().startswith(
        (*_FORMULA_PREFIXES, "-")
    )


def _date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and 1 <= float(value) <= 2958465:
        # Excel's 1900 epoch including its historical leap-year bug.
        return date(1899, 12, 30) + timedelta(days=int(float(value)))
    if not isinstance(value, str):
        return None
    clean = value.strip()
    for pattern in ("%Y.%m.%d", "%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(clean, pattern).date()
        except ValueError:
            continue
    return None


def _fire_values(
    value: Any,
) -> tuple[int | None, int | float | None, str | None]:
    if value is None or value == "":
        return None, None, None
    if isinstance(value, (int, float)):
        return (
            None,
            None,
            "火工品量是包含不同单位子项的类别，单个数值无法判断是雷管还是炸药",
        )
    raw_text = str(value).strip()
    if _number(raw_text) is not None:
        return (
            None,
            None,
            "火工品量是包含不同单位子项的类别，单个数值无法判断是雷管还是炸药",
        )
    matches: dict[str, list[str]] = {"雷管": [], "炸药": []}
    for name, raw in _FIRE_NUMBER.findall(raw_text):
        matches[name].append(raw)
    if not matches["雷管"] and not matches["炸药"]:
        return None, None, "火工品量未明确写出雷管和/或炸药子项"

    warnings: list[str] = []
    detonators: int | None = None
    explosives: int | float | None = None
    if len(matches["雷管"]) > 1:
        warnings.append("同一单元格包含多个雷管数值，无法判断是否已含合计")
    elif matches["雷管"]:
        detonators = _number(matches["雷管"][0], integer=True)
        if detonators is None:
            warnings.append("雷管数量必须是非负整数")
    if len(matches["炸药"]) > 1:
        warnings.append("同一单元格包含多个炸药数值，无法判断是否已含合计")
    elif matches["炸药"]:
        explosives = _number(matches["炸药"][0])
        if explosives is None:
            warnings.append("炸药数量必须是非负数")

    residual = _FIRE_NUMBER.sub("", raw_text)
    residual = re.sub(
        r"(?:千克|公斤|kg|KG|发|枚|个|\s|[,，;；、/|()（）+])+",
        "",
        residual,
    )
    if residual:
        warnings.append(f"还包含未识别的火工品内容：{residual[:40]}")
    return detonators, explosives, "；".join(warnings) or None


def _measurement(metric: str, value: Any, source_id: str) -> dict[str, Any]:
    parsed = _metric_number(metric, value)
    return {
        "metric_code": metric,
        "value": parsed,
        "unit": UNITS[metric],
        "aggregation": AGGREGATIONS[metric],
        "quality_flags": ["reported"] if parsed is not None else ["missing"],
        "source_refs": [source_id],
    }


def _empty_set(source_id: str) -> dict[str, dict[str, Any]]:
    return {metric: _measurement(metric, None, source_id) for metric in METRICS}


def _shift_window(day: date, key: str, timezone: str) -> tuple[str, str]:
    zone = ZoneInfo(timezone)
    start_hours = {"zero_shift": 0, "eight_shift": 8, "four_shift": 16}
    start = datetime.combine(day, time(start_hours[key]), tzinfo=zone)
    end = start + timedelta(hours=8)
    return start.isoformat(), end.isoformat()


def _source(
    *,
    source_id: str,
    acquisition_mode: str,
    filename: str,
    sheet: str,
    row_number: int,
    captured_at: str,
    content_hash: str,
    media_type: str,
) -> dict[str, Any]:
    return {
        "source_id": source_id,
        "acquisition_mode": acquisition_mode,
        "source_system": "enterprise-five-quantity-import",
        "source_record_id": f"{filename}#{sheet}!row-{row_number}",
        "source_location": f"{sheet}!row-{row_number}",
        "captured_at": captured_at,
        "media_type": media_type,
        "evidence_sha256": content_hash,
        "normalization": (
            "Locally validated header mapping to V2 fixed units; missing values "
            "remain null and no value is estimated or imputed."
        ),
    }


def _metric_match(value: str) -> tuple[str, bool, str] | None:
    for alias, metric, legacy in _METRIC_ALIASES:
        if alias in value:
            return metric, legacy, alias
    return None


def _shift_match(value: str) -> str | None:
    return next(
        (shift for alias, shift in _SHIFT_ALIASES.items() if alias in value),
        None,
    )


def csv_header_unit_issue(metric: str, source_header: str) -> str | None:
    """Reject explicit source units that cannot be silently treated as V2 units."""

    if metric not in METRICS or not isinstance(source_header, str):
        return "来源字段或规范指标非法"
    normal = _normal_text(source_header)
    if metric == "production_t" and any(
        token in normal for token in ("万吨", "千吨", "公斤", "千克", "(kg)", "（kg）")
    ):
        return "来源产量单位不是吨；系统不会静默换算，请先按吨导出"
    if metric == "explosives_kg" and any(
        token in normal for token in ("万吨", "千吨", "(t)", "（t）", "吨")
    ):
        return "来源炸药单位不是千克；系统不会静默换算，请先按千克导出"
    if metric == "electricity_kwh" and any(
        token in normal
        for token in ("mwh", "兆瓦时", "万度", "(wh)", "（wh）")
    ):
        return "来源电量单位不是千瓦时；系统不会静默换算，请先按 kWh 导出"
    if metric == "ventilation_m3_min" and any(
        token in normal
        for token in (
            "m3/s",
            "m³/s",
            "m3/h",
            "m³/h",
            "立方米/秒",
            "立方米每秒",
            "立方米/小时",
            "立方米每小时",
        )
    ):
        return "来源风量单位不是立方米/分钟；系统不会静默换算，请先按 m3/min 导出"
    return None


def _looks_like_header_detail(row: list[Any], date_column: int) -> bool:
    for index, value in enumerate(row):
        if index == date_column or not isinstance(value, str):
            continue
        normal = _normal_text(value)
        if normal and (_metric_match(normal) or _shift_match(normal)):
            return True
    return False


def _filled_header_row(row: list[Any], width: int) -> list[str]:
    result: list[str] = []
    inherited = ""
    for column in range(width):
        value = _normal_text(row[column]) if column < len(row) else ""
        if value:
            inherited = value
        result.append(value or inherited)
    return result


def _display_filled_header_row(row: list[Any], width: int) -> list[str]:
    result: list[str] = []
    inherited = ""
    for column in range(width):
        raw = row[column] if column < len(row) else ""
        value = str(raw).strip() if raw is not None else ""
        if value:
            inherited = value
        result.append(value or inherited)
    return result


def _build_layout(
    sheet: ParsedSheet,
    *,
    header_start: int,
    data_start: int,
    date_column: int,
    date_inferred: bool,
) -> TableLayout:
    header_rows = sheet.rows[header_start:data_start]
    if not header_rows:
        raise ImportContentError("日期数据之前没有可读取的表头")
    width = min(MAX_COLUMNS, max(len(header) for header in header_rows))
    normalized_rows = [
        _filled_header_row(header, width) if len(header_rows) > 1 else [
            _normal_text(header[column]) if column < len(header) else ""
            for column in range(width)
        ]
        for header in header_rows
    ]
    display_rows = [
        _display_filled_header_row(header, width) if len(header_rows) > 1 else [
            str(header[column]).strip()
            if column < len(header) and header[column] is not None
            else ""
            for column in range(width)
        ]
        for header in header_rows
    ]
    normalized_headers = tuple(
        "|".join(row[column] for row in normalized_rows if row[column])
        for column in range(width)
    )
    display_headers = tuple(
        " / ".join(
            dict.fromkeys(row[column] for row in display_rows if row[column])
        )[:240]
        for column in range(width)
    )
    return TableLayout(
        header_start=header_start,
        data_start=data_start,
        date_column=date_column,
        width=width,
        normalized_headers=normalized_headers,
        display_headers=display_headers,
        date_inferred=date_inferred,
    )


def _safe_inferred_header_row(row: list[Any], width: int) -> bool:
    """Return true only when a fallback header is clearly metadata, not data.

    Date-column inference is allowed for opaque ERP labels such as ``业务日``,
    but it must never promote the first observation row to a header.  Besides
    losing that observation, doing so could expose business values to the
    optional header-mapping model.  Numeric, date-like and formula-like cells
    therefore make an inferred header fail closed.
    """

    labels = 0
    for column in range(width):
        value = row[column] if column < len(row) else None
        if _explicit_missing(value):
            continue
        if (
            _date_value(value) is not None
            or _number(value) is not None
            or _formula_like(value)
        ):
            return False
        text = str(value).strip()
        if not _normal_text(value) or _HEADER_NUMBER_WITH_UNIT.fullmatch(text):
            return False
        labels += 1
    return labels >= 2


def _safe_declared_header_row(
    row: list[Any],
    *,
    minimum_labels: int = 2,
) -> bool:
    """Reject a declared header row that contains observation-like cells."""

    labels = 0
    for value in row[:MAX_COLUMNS]:
        if _explicit_missing(value):
            continue
        if (
            _date_value(value) is not None
            or _number(value) is not None
            or _formula_like(value)
            or _HEADER_NUMBER_WITH_UNIT.fullmatch(str(value).strip())
        ):
            return False
        if not _normal_text(value):
            return False
        labels += 1
    return labels >= minimum_labels


def _safe_inferred_date_header(value: Any) -> bool:
    normal = _normal_text(value)
    return bool(normal) and any(
        hint in normal for hint in _INFERRED_DATE_HEADER_HINTS
    )


def _find_table_layout(sheet: ParsedSheet) -> TableLayout:
    rows = sheet.rows
    for date_header_row in range(min(12, len(rows))):
        row = rows[date_header_row]
        date_column = next(
            (
                index
                for index, value in enumerate(row)
                if _normal_text(value) in _DATE_ALIASES
            ),
            None,
        )
        if date_column is None:
            continue
        if not _safe_declared_header_row(row):
            continue
        data_start = date_header_row + 1
        while data_start < min(len(rows), date_header_row + 3):
            candidate = rows[data_start]
            candidate_date = (
                candidate[date_column] if date_column < len(candidate) else None
            )
            if _date_value(candidate_date) is not None:
                break
            if not _looks_like_header_detail(candidate, date_column):
                break
            if not _safe_declared_header_row(
                candidate,
                minimum_labels=1,
            ):
                raise ImportContentError(
                    "表头明细行包含疑似观测值、日期或公式，无法安全预览"
                )
            data_start += 1
        return _build_layout(
            sheet,
            header_start=date_header_row,
            data_start=data_start,
            date_column=date_column,
            date_inferred=False,
        )

    # Some enterprise exports use unlisted but recognisable labels such as
    # "业务日".  Infer the date column only when a clearly textual, date-labelled
    # header precedes one unambiguous column of parseable calendar dates.  This
    # conservative boundary prevents a headerless first data row from becoming
    # model-visible metadata.
    candidates: list[tuple[float, int, int, int]] = []
    search_end = min(len(rows), 16)
    for first_data_row in range(1, search_end):
        width = min(MAX_COLUMNS, max(len(row) for row in rows[:search_end]))
        header_start = first_data_row - 1
        if not _safe_inferred_header_row(rows[header_start], width):
            continue
        for column in range(width):
            header_value = (
                rows[header_start][column]
                if column < len(rows[header_start])
                else None
            )
            if not _safe_inferred_date_header(header_value):
                continue
            if any(
                _date_value(
                    rows[previous][column]
                    if column < len(rows[previous])
                    else None
                )
                is not None
                for previous in range(header_start)
            ):
                continue
            first_value = (
                rows[first_data_row][column]
                if column < len(rows[first_data_row])
                else None
            )
            if _date_value(first_value) is None:
                continue
            non_empty = 0
            parsed = 0
            for candidate_row in rows[first_data_row : first_data_row + 20]:
                value = candidate_row[column] if column < len(candidate_row) else None
                if value in {None, ""}:
                    continue
                non_empty += 1
                if _date_value(value) is not None:
                    parsed += 1
            if parsed and non_empty and parsed / non_empty >= 0.8:
                candidates.append((parsed / non_empty, parsed, first_data_row, column))
    if candidates:
        candidates.sort(reverse=True)
        best = candidates[0]
        competing_columns = {
            item[3]
            for item in candidates
            if item[:2] == best[:2] and item[2] == best[2]
        }
        if len(competing_columns) == 1:
            first_data_row = best[2]
            date_column = best[3]
            header_start = first_data_row - 1
            return _build_layout(
                sheet,
                header_start=header_start,
                data_start=first_data_row,
                date_column=date_column,
                date_inferred=True,
            )
    if not rows:
        raise ImportContentError("文件中没有可读取的表格")
    raise ImportContentError("未找到可安全确定的日期列")


def _find_table(
    sheet: ParsedSheet,
    *,
    mapping_override: dict[int, tuple[str, str]] | None = None,
) -> tuple[int, int, dict[int, tuple[str, str]], list[dict[str, Any]]]:
    layout = _find_table_layout(sheet)
    mapping: dict[int, tuple[str, str]] = {}
    warnings: list[dict[str, Any]] = []
    if layout.date_inferred:
        warnings.append(
            {
                "kind": "inferred_date_column",
                "source_column": layout.date_column,
                "source_header": layout.display_headers[layout.date_column],
                "reason": "日期列由单元格日期格式推断，生成草稿前须人工核对",
                "requires_human_review": True,
            }
        )
    for column in range(layout.width):
        if column == layout.date_column:
            continue
        combined = layout.normalized_headers[column]
        if mapping_override is not None:
            target = mapping_override.get(column)
            if target is not None:
                if target[0] != "fire_material":
                    unit_issue = csv_header_unit_issue(
                        target[0], layout.display_headers[column]
                    )
                    if unit_issue is not None:
                        raise ImportContentError(unit_issue)
                mapping[column] = target
            elif combined:
                warnings.append(
                    {
                        "kind": "explicitly_unmapped_column",
                        "source_column": column,
                        "source_header": layout.display_headers[column],
                        "reason": "该来源列未被人工映射，未写入草稿",
                        "requires_human_review": True,
                    }
                )
            continue
        matched = _metric_match(combined)
        if matched is None:
            unsupported = next(
                (
                    alias
                    for alias in _UNSUPPORTED_FIRE_COMPONENT_ALIASES
                    if alias in combined
                ),
                None,
            )
            if unsupported:
                warnings.append(
                    {
                        "kind": "unsupported_fire_material_component",
                        "source_column": column,
                        "source_alias": unsupported,
                        "reason": (
                            f"火工品子项“{unsupported}”尚无已批准的规范单位和"
                            "报送字段，未自动归入雷管或炸药"
                        ),
                        "requires_human_review": True,
                    }
                )
            elif combined:
                warnings.append(
                    {
                        "kind": "unmapped_column",
                        "source_column": column,
                        "source_header": layout.display_headers[column],
                        "reason": (
                            "来源列未映射到日期、五量或班次字段，未自动写入草稿："
                            f"{layout.display_headers[column][:80]}"
                        ),
                        "requires_human_review": True,
                    }
                )
            continue
        metric, legacy, alias = matched
        shift = _shift_match(combined) or "daily_total"
        if metric != "fire_material":
            unit_issue = csv_header_unit_issue(
                metric,
                layout.display_headers[column],
            )
            if unit_issue is not None:
                warnings.append(
                    {
                        "kind": "incompatible_source_unit",
                        "source_column": column,
                        "source_header": layout.display_headers[column],
                        "reason": unit_issue,
                        "requires_human_review": True,
                    }
                )
                continue
        mapping[column] = (metric, shift)
        if legacy:
            warnings.append(
                {
                    "kind": "legacy_mine_entry_alias",
                    "source_column": column,
                    "source_alias": alias,
                    "reason": (
                        "旧字段已按入井人员量解析；请对照原始口径确认其表示入井人数/人次"
                    ),
                    "requires_human_review": True,
                }
            )
    if mapping:
        return layout.data_start, layout.date_column, mapping, warnings
    if mapping_override is not None:
        raise ImportContentError("至少需要确认一个五量字段映射")
    raise ImportContentError("未找到可识别的五量表头")


def _normalise_sheet(
    sheet: ParsedSheet,
    *,
    filename: str,
    acquisition_mode: str,
    content_hash: str,
    captured_at: str,
    media_type: str,
    timezone: str,
    mapping_override: dict[int, tuple[str, str]] | None = None,
    mapping_source: str = "deterministic",
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    start_row, date_column, mapping, header_warnings = _find_table(
        sheet,
        mapping_override=mapping_override,
    )
    days: list[dict[str, Any]] = []
    sources: list[dict[str, Any]] = []
    suggestions = header_warnings + [
        {
            "kind": "column_mapping",
            "source_column": column,
            "metric": metric,
            "period": period,
            "confidence": 1.0,
            "source": mapping_source,
            "requires_human_review": True,
        }
        for column, (metric, period) in sorted(mapping.items())
    ]
    columns_by_target: dict[tuple[str, str], list[int]] = {}
    for column, target in mapping.items():
        columns_by_target.setdefault(target, []).append(column)
    blocked_columns: set[int] = set()
    for (metric, period), columns in columns_by_target.items():
        if len(columns) <= 1:
            continue
        blocked_columns.update(columns)
        suggestions.append(
            {
                "kind": "duplicate_column_mapping",
                "source_columns": columns,
                "metric": metric,
                "period": period,
                "reason": (
                    "多个来源列指向同一数据项，无法判断其中是否已含合计，"
                    "未自动覆盖或相加"
                ),
                "requires_human_review": True,
            }
        )
    for period in ("daily_total", *SHIFT_KEYS):
        generic_columns = columns_by_target.get(("fire_material", period), [])
        explicit_columns = [
            *columns_by_target.get(("detonators_count", period), []),
            *columns_by_target.get(("explosives_kg", period), []),
        ]
        if generic_columns and explicit_columns:
            blocked_columns.update(generic_columns)
            suggestions.append(
                {
                    "kind": "overlapping_fire_material_columns",
                    "source_columns": generic_columns + explicit_columns,
                    "period": period,
                    "reason": (
                        "同时存在火工品总栏和明确子项栏；为避免重复计算，"
                        "仅采用明确的雷管/炸药子项栏"
                    ),
                    "requires_human_review": True,
                }
            )
    for row_index in range(start_row, min(len(sheet.rows), MAX_ROWS)):
        row = sheet.rows[row_index]
        raw_date = row[date_column] if date_column < len(row) else None
        day_value = _date_value(raw_date)
        if day_value is None:
            if any(value not in {None, ""} for value in row):
                suggestions.append(
                    {
                        "kind": "skipped_row",
                        "source_row": row_index + 1,
                        "reason": "日期无法确定，未猜测",
                        "requires_human_review": True,
                    }
                )
            continue
        source_id = f"SRC-{content_hash[:12]}-{row_index + 1:04d}"
        source = _source(
            source_id=source_id,
            acquisition_mode=acquisition_mode,
            filename=filename,
            sheet=sheet.name,
            row_number=row_index + 1,
            captured_at=captured_at,
            content_hash=content_hash,
            media_type=media_type,
        )
        sources.append(source)
        daily = _empty_set(source_id)
        shifts = {key: _empty_set(source_id) for key in SHIFT_KEYS}
        for column, (group, period) in mapping.items():
            if column in blocked_columns:
                continue
            value = row[column] if column < len(row) else None
            target = daily if period == "daily_total" else shifts[period]
            if group == "fire_material":
                detonators, explosives, warning = _fire_values(value)
                target["detonators_count"] = _measurement(
                    "detonators_count", detonators, source_id
                )
                target["explosives_kg"] = _measurement(
                    "explosives_kg", explosives, source_id
                )
                if warning:
                    suggestions.append(
                        {
                            "kind": "ambiguous_fire_material_value",
                            "source_column": column,
                            "source_row": row_index + 1,
                            "period": period,
                            "reason": warning,
                            "requires_human_review": True,
                        }
                    )
            else:
                measurement = _measurement(group, value, source_id)
                target[group] = measurement
                if not _explicit_missing(value) and measurement["value"] is None:
                    formula_like = _formula_like(value)
                    measurement["quality_flags"].append("source_format_warning")
                    suggestions.append(
                        {
                            "kind": (
                                "formula_like_cell"
                                if formula_like
                                else "invalid_numeric_cell"
                            ),
                            "source_column": column,
                            "source_row": row_index + 1,
                            "metric": group,
                            "period": period,
                            "reason": (
                                "单元格疑似公式、命令前缀或不允许的负数，"
                                "未执行且未写入数值"
                                if formula_like
                                else "单元格不是可安全确定的非负数，未猜测或写入数值"
                            ),
                            "requires_human_review": True,
                        }
                    )
        shift_documents: dict[str, Any] = {}
        for key in SHIFT_KEYS:
            start_at, end_at = _shift_window(day_value, key, timezone)
            shift_documents[key] = {
                "shift_code": {
                    "zero_shift": "ZERO",
                    "eight_shift": "EIGHT",
                    "four_shift": "FOUR",
                }[key],
                "start_at": start_at,
                "end_at": end_at,
                "measurements": shifts[key],
            }
        production = daily["production_t"]["value"]
        operating_state = (
            "unknown"
            if production is None
            else "stopped"
            if float(production) == 0
            else "producing"
        )
        days.append(
            {
                "date": day_value.isoformat(),
                "operating_state": operating_state,
                "reported_quantity": {
                    "daily_total": daily,
                    "shifts": shift_documents,
                },
            }
        )
    return days, sources, suggestions


def _json_payload(
    content: bytes,
    *,
    filename: str,
    acquisition_mode: str,
    content_hash: str,
    captured_at: str,
    identity: MineIdentity,
    media_type: str = "application/json",
    source_sheet: str = "JSON",
) -> dict[str, Any] | None:
    try:
        parsed = _strict_json_loads(_decode_text(content))
    except json.JSONDecodeError as error:
        raise ImportContentError("JSON 文件格式非法") from error
    candidate = parsed.get("payload") if isinstance(parsed, dict) else None
    if candidate is None and isinstance(parsed, dict) and "days" in parsed:
        candidate = parsed
    if not isinstance(candidate, dict) or not isinstance(candidate.get("days"), list):
        return None
    days = candidate["days"]
    if not days:
        raise ImportContentError("JSON days 不能为空")
    # Do not accept identity, confirmation, signatures or provenance authority
    # from an imported file.  Those are rebuilt under this one-mine instance.
    source_id = f"SRC-{content_hash[:16]}"
    source = _source(
        source_id=source_id,
        acquisition_mode=acquisition_mode,
        filename=filename,
        sheet=source_sheet,
        row_number=1,
        captured_at=captured_at,
        content_hash=content_hash,
        media_type=media_type,
    )
    clean_days: list[dict[str, Any]] = []
    for index, item in enumerate(days):
        if not isinstance(item, dict):
            raise ImportContentError(f"JSON days[{index}] 必须是对象")
        day_value = _date_value(item.get("date"))
        if day_value is None:
            raise ImportContentError(f"JSON days[{index}].date 非法")
        quantity = item.get("reported_quantity")
        if not isinstance(quantity, dict):
            raise ImportContentError(f"JSON days[{index}] 缺少 reported_quantity")
        # Round-trip through the strict local validator later.  Replace all
        # source refs so the imported document cannot claim another authority.
        copied = json.loads(jcs_json(item))
        for measurement_set in [copied["reported_quantity"]["daily_total"]] + [
            copied["reported_quantity"]["shifts"][key]["measurements"]
            for key in SHIFT_KEYS
        ]:
            if "labor_persons" in measurement_set:
                if "mine_entry_persons" in measurement_set:
                    raise ImportContentError(
                        "JSON 同时包含 mine_entry_persons 和旧 labor_persons，"
                        "无法判断是否重复"
                    )
                legacy = measurement_set.pop("labor_persons")
                if not isinstance(legacy, dict):
                    raise ImportContentError("JSON 旧 labor_persons 测量结构非法")
                legacy["metric_code"] = "mine_entry_persons"
                measurement_set["mine_entry_persons"] = legacy
            for metric in METRICS:
                if metric not in measurement_set:
                    raise ImportContentError(f"JSON 缺少规范数据项 {metric}")
                measurement_set[metric]["source_refs"] = [source_id]
        clean_days.append(copied)
    clean_days.sort(key=lambda item: item["date"])
    return _draft_payload(
        identity=identity,
        days=clean_days,
        sources=[source],
        captured_at=captured_at,
        suggestions=[
            {
                "kind": "structured_json_mapping",
                "confidence": 1.0,
                "requires_human_review": True,
            }
        ],
    )


def _jsonl_payload(
    content: bytes,
    *,
    filename: str,
    acquisition_mode: str,
    content_hash: str,
    captured_at: str,
    identity: MineIdentity,
) -> dict[str, Any]:
    days: list[dict[str, Any]] = []
    for line_number, raw_line in enumerate(_decode_text(content).splitlines(), start=1):
        line = raw_line.strip()
        if not line:
            continue
        if line_number > MAX_ROWS:
            raise ImportContentError(f"JSONL 行数不得超过 {MAX_ROWS}")
        try:
            item = _strict_json_loads(line)
        except json.JSONDecodeError as error:
            raise ImportContentError(
                f"JSONL 第 {line_number} 行不是有效 JSON"
            ) from error
        if not isinstance(item, dict):
            raise ImportContentError(f"JSONL 第 {line_number} 行必须是对象")
        candidate = (
            item.get("payload") if isinstance(item.get("payload"), dict) else item
        )
        if isinstance(candidate.get("days"), list):
            days.extend(candidate["days"])
        elif "date" in candidate and "reported_quantity" in candidate:
            days.append(candidate)
        else:
            raise ImportContentError(
                f"JSONL 第 {line_number} 行必须是日报，或包含 payload.days/days"
            )
        if len(days) > 366:
            raise ImportContentError("JSONL 一次最多导入 366 个日报")
    if not days:
        raise ImportContentError("JSONL 没有可读取的日报")
    payload = _json_payload(
        json.dumps({"days": days}, ensure_ascii=False).encode("utf-8"),
        filename=filename,
        acquisition_mode=acquisition_mode,
        content_hash=content_hash,
        captured_at=captured_at,
        identity=identity,
        media_type="application/x-ndjson",
        source_sheet="JSONL",
    )
    if payload is None:
        raise ImportContentError("JSONL 日报结构非法")
    return payload


def _draft_payload(
    *,
    identity: MineIdentity,
    days: list[dict[str, Any]],
    sources: list[dict[str, Any]],
    captured_at: str,
    suggestions: list[dict[str, Any]],
    model_assistance_used: bool = False,
    model_output_sha256: str | None = None,
    processing_rule: str = "five-quantity-deterministic-normalizer-v2",
) -> dict[str, Any]:
    if not days:
        raise ImportContentError("未提取到任何有效日报")
    days.sort(key=lambda item: item["date"])
    dates = [item["date"] for item in days]
    if len(dates) != len(set(dates)):
        raise ImportContentError("文件中存在重复日期，请人工处理后重新导入")
    months = {value[:7] for value in dates}
    if len(months) != 1:
        raise ImportContentError("一次导入只能形成一个月份的报表")
    processing_record = {
        "captured_at": captured_at,
        "suggestions": suggestions,
        "rule": processing_rule,
        "model_output_sha256": model_output_sha256,
    }
    agent_processing: dict[str, Any] = {
        "normalization_performed": True,
        "model_assistance_used": model_assistance_used,
        "processing_record_sha256": hashlib.sha256(
            jcs_json(processing_record).encode("utf-8")
        ).hexdigest(),
    }
    if model_assistance_used:
        if (
            not isinstance(model_output_sha256, str)
            or len(model_output_sha256) != 64
            or any(
                character not in "0123456789abcdef"
                for character in model_output_sha256
            )
        ):
            raise ImportContentError("模型参与字段映射时必须记录模型输出 SHA-256")
        agent_processing["model_output_sha256"] = model_output_sha256
    elif model_output_sha256 is not None:
        raise ImportContentError("模型未参与字段映射时不得声明模型输出摘要")
    return {
        "mine": identity.mine,
        "reporting_month": next(iter(months)),
        "timezone": identity.timezone,
        "period_start": min(dates),
        "period_end": max(dates),
        "closed_at": captured_at,
        "comparison_context": identity.comparison_context,
        "days": days,
        "sources": sources,
        "agent_processing": agent_processing,
    }


def _masked_value_type(value: Any) -> str:
    if _explicit_missing(value):
        return "empty"
    if _formula_like(value):
        return "unsafe_formula_like"
    if _date_value(value) is not None:
        return "date"
    parsed = _number(value)
    if parsed is not None:
        return "integer" if isinstance(parsed, int) else "decimal"
    return "text"


def inspect_five_quantity_csv(
    *,
    filename: str,
    content: bytes,
) -> dict[str, Any]:
    """Inspect a CSV without creating a draft or exposing its raw values.

    The result contains source headers and bounded value *types* only.  It is
    therefore safe to pass to a constrained mapping assistant; actual business
    values stay inside the local parser and are never part of the model prompt.
    """

    name = _check_bytes(filename, content)
    if not name.casefold().endswith(".csv"):
        raise ImportContentError("智能字段映射预览目前仅支持 CSV")
    rows, encoding, delimiter = _csv_rows(content)
    sheet = ParsedSheet("CSV", rows)
    layout = _find_table_layout(sheet)
    valid_dates: list[date] = []
    row_count = 0
    for row in rows[layout.data_start:]:
        if any(value not in {None, ""} for value in row):
            row_count += 1
        raw_date = row[layout.date_column] if layout.date_column < len(row) else None
        parsed_date = _date_value(raw_date)
        if parsed_date is not None:
            valid_dates.append(parsed_date)

    columns: list[dict[str, Any]] = []
    for column in range(layout.width):
        if column == layout.date_column:
            continue
        values = [
            row[column] if column < len(row) else None
            for row in rows[layout.data_start : layout.data_start + 50]
        ]
        non_empty_count = sum(not _explicit_missing(value) for value in values)
        header = layout.display_headers[column] or f"第 {column + 1} 列（无表头）"
        normalized = layout.normalized_headers[column]
        if not normalized and non_empty_count == 0:
            continue
        type_counts: dict[str, int] = {}
        for value in values:
            kind = _masked_value_type(value)
            type_counts[kind] = type_counts.get(kind, 0) + 1
        target_metric: str | None = None
        target_period: str | None = None
        target_unit: str | None = None
        confidence = 0.0
        source = "deterministic"
        status = "unmapped"
        reason = "本地规则尚不能确定该列含义，请人工选择或明确忽略"
        matched = _metric_match(normalized)
        unsupported = next(
            (
                alias
                for alias in _UNSUPPORTED_FIRE_COMPONENT_ALIASES
                if alias in normalized
            ),
            None,
        )
        if matched is not None:
            target_metric, legacy, alias = matched
            if target_metric == "fire_material":
                target_metric = None
                status = "blocked"
                reason = (
                    "火工品总栏同时包含不同单位；请拆成雷管和炸药两列，"
                    "或人工确认该列只代表其中一项"
                )
            else:
                target_period = _shift_match(normalized) or "daily_total"
                target_unit = UNITS[target_metric]
                confidence = 0.82 if legacy else 0.99
                status = "needs_review" if legacy else "mapped"
                reason = (
                    "旧字段别名可能存在口径差异，请确认其确为入井人数/人次"
                    if legacy
                    else f"表头包含已批准别名“{alias}”"
                )
                unit_issue = csv_header_unit_issue(target_metric, header)
                if unit_issue is not None:
                    target_metric = None
                    target_period = None
                    target_unit = None
                    confidence = 0.0
                    status = "blocked"
                    reason = unit_issue
        elif unsupported is not None:
            status = "blocked"
            reason = f"“{unsupported}”没有已批准的五量目标字段，不能自动归类"
        columns.append(
            {
                "source_index": column,
                "source_header": header[:240],
                "normalized_header": normalized[:240],
                "target_metric": target_metric,
                "target_period": target_period,
                "target_unit": target_unit,
                "confidence": confidence,
                "source": source,
                "reason": reason,
                "status": status,
                "sample_types": type_counts,
                "non_empty_sample_count": non_empty_count,
            }
        )

    columns_by_target: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for item in columns:
        metric = item["target_metric"]
        period = item["target_period"]
        if metric is not None and period is not None:
            columns_by_target.setdefault((metric, period), []).append(item)
    warnings: list[dict[str, Any]] = []
    for (metric, period), items in columns_by_target.items():
        if len(items) <= 1:
            continue
        for item in items:
            item["status"] = "blocked"
            item["reason"] = "多个来源列指向同一规范字段，必须人工保留其中一列"
        warnings.append(
            {
                "code": "duplicate_target",
                "message": "多个来源列被识别为同一规范字段，请人工选择",
                "severity": "error",
                "source_indexes": [item["source_index"] for item in items],
                "target_metric": metric,
                "target_period": period,
            }
        )
    if layout.date_inferred:
        warnings.append(
            {
                "code": "inferred_date_column",
                "message": "日期列根据单元格格式推断，请在生成草稿后重点核对日期",
                "severity": "warning",
                "source_index": layout.date_column,
            }
        )
    content_hash = hashlib.sha256(content).hexdigest()
    schema_fingerprint = hashlib.sha256(
        jcs_json(
            {
                "date_column": layout.date_column,
                "headers": list(layout.normalized_headers),
            }
        ).encode("utf-8")
    ).hexdigest()
    date_header = layout.display_headers[layout.date_column] or (
        f"第 {layout.date_column + 1} 列"
    )
    return {
        "contract_version": "five-quantity-csv-inspection/v1",
        "content_sha256": content_hash,
        "schema_fingerprint": schema_fingerprint,
        "filename": name,
        "byte_size": len(content),
        "encoding": encoding,
        "delimiter": delimiter,
        "header_row": layout.header_start + 1,
        "data_start_row": layout.data_start + 1,
        "date_column": {
            "source_index": layout.date_column,
            "source_header": date_header[:240],
            "inferred": layout.date_inferred,
            "confidence": 0.85 if layout.date_inferred else 1.0,
        },
        "columns": columns,
        "row_count": row_count,
        "valid_day_count": len(valid_dates),
        "detected_months": sorted({value.strftime("%Y-%m") for value in valid_dates}),
        "warnings": warnings,
    }


def _column_mapping_override(
    mappings: list[dict[str, Any]] | None,
) -> dict[int, tuple[str, str]] | None:
    if mappings is None:
        return None
    if not isinstance(mappings, list) or not 1 <= len(mappings) <= MAX_COLUMNS:
        raise ImportContentError("CSV 字段映射必须包含 1-256 个来源列")
    result: dict[int, tuple[str, str]] = {}
    targets: set[tuple[str, str]] = set()
    for item in mappings:
        if not isinstance(item, dict) or set(item) != {
            "source_index",
            "target_metric",
            "target_period",
        }:
            raise ImportContentError("CSV 字段映射结构非法")
        column = item["source_index"]
        metric = item["target_metric"]
        period = item["target_period"]
        if (
            isinstance(column, bool)
            or not isinstance(column, int)
            or not 0 <= column < MAX_COLUMNS
        ):
            raise ImportContentError("CSV 来源列编号非法")
        if metric not in {*METRICS, "fire_material"} or period not in PERIOD_KEYS:
            raise ImportContentError("CSV 字段映射目标不在五量白名单内")
        target = (metric, period)
        if column in result:
            raise ImportContentError("同一 CSV 来源列不能重复映射")
        if target in targets:
            raise ImportContentError("多个 CSV 来源列不能映射到同一规范字段")
        result[column] = target
        targets.add(target)
    return result


def import_five_quantity_bytes(
    *,
    filename: str,
    content: bytes,
    acquisition_mode: str,
    identity: MineIdentity,
    captured_at: str | None = None,
    column_mappings: list[dict[str, Any]] | None = None,
    model_assistance_used: bool = False,
    model_output_sha256: str | None = None,
) -> dict[str, Any]:
    """Parse an input without inventing values and return a reviewable draft."""

    name = _check_bytes(filename, content)
    if not isinstance(model_assistance_used, bool):
        raise ImportContentError("model_assistance_used 必须是布尔值")
    if acquisition_mode not in {"manual_import", "direct_collection"}:
        raise ImportContentError("采集方式只能是 manual_import 或 direct_collection")
    captured = captured_at or utc_text()
    content_hash = hashlib.sha256(content).hexdigest()
    suffix = "." + name.rsplit(".", 1)[-1].lower()
    if column_mappings is not None and suffix != ".csv":
        raise ImportContentError("显式字段映射目前仅适用于 CSV")
    mapping_override = _column_mapping_override(column_mappings)
    if suffix == ".jsonl":
        structured = _jsonl_payload(
            content,
            filename=name,
            acquisition_mode=acquisition_mode,
            content_hash=content_hash,
            captured_at=captured,
            identity=identity,
        )
        return {
            "content_sha256": content_hash,
            "filename": name,
            "acquisition_mode": acquisition_mode,
            "payload": structured,
            "suggestions": [
                {
                    "kind": "structured_jsonl_mapping",
                    "confidence": 1.0,
                    "requires_human_review": True,
                }
            ],
        }
    if suffix == ".json":
        structured = _json_payload(
            content,
            filename=name,
            acquisition_mode=acquisition_mode,
            content_hash=content_hash,
            captured_at=captured,
            identity=identity,
        )
        if structured is not None:
            return {
                "content_sha256": content_hash,
                "filename": name,
                "acquisition_mode": acquisition_mode,
                "payload": structured,
                "suggestions": [
                    {
                        "kind": "structured_json_mapping",
                        "confidence": 1.0,
                        "requires_human_review": True,
                    }
                ],
            }
        raise ImportContentError("JSON 必须包含 V2 payload.days 或 days")
    if suffix == ".csv":
        sheets = _csv_sheets(content)
        media_type = "text/csv"
    elif content.startswith(b"PK\x03\x04"):
        sheets = _xlsx_sheets(content)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif content.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        sheets = _xls_sheets(content)
        media_type = "application/vnd.ms-excel"
    else:
        raise ImportContentError("工作簿内容与 ET/XLS/XLSX 扩展名不匹配")
    all_days: list[dict[str, Any]] = []
    all_sources: list[dict[str, Any]] = []
    all_suggestions: list[dict[str, Any]] = []
    errors: list[str] = []
    for sheet in sheets:
        try:
            days, sources, suggestions = _normalise_sheet(
                sheet,
                filename=name,
                acquisition_mode=acquisition_mode,
                content_hash=content_hash,
                captured_at=captured,
                media_type=media_type,
                timezone=identity.timezone,
                mapping_override=mapping_override,
                mapping_source=(
                    "human_reviewed_preview"
                    if mapping_override is not None
                    else "deterministic"
                ),
            )
        except ImportContentError as error:
            errors.append(f"{sheet.name}: {error}")
            continue
        if days:
            all_days.extend(days)
            all_sources.extend(sources)
            all_suggestions.extend(suggestions)
    if not all_days:
        detail = "；".join(errors[:5]) if errors else "没有包含五量数据的 sheet"
        raise ImportContentError(f"未能规范化工作簿：{detail}")
    payload = _draft_payload(
        identity=identity,
        days=all_days,
        sources=all_sources,
        captured_at=captured,
        suggestions=all_suggestions,
        model_assistance_used=model_assistance_used,
        model_output_sha256=model_output_sha256,
        processing_rule=(
            "five-quantity-reviewed-csv-mapping-v1"
            if mapping_override is not None
            else "five-quantity-deterministic-normalizer-v2"
        ),
    )
    return {
        "content_sha256": content_hash,
        "filename": name,
        "acquisition_mode": acquisition_mode,
        "payload": payload,
        "suggestions": all_suggestions,
    }
